"""
每日条件单Excel生成模块
========================
盘前自动生成条件单清单，包含:
- 买入条件单价格、买入数量
- 止损条件单价格（收盘价触发）
- 止盈条件单价格
输出为Excel表格，打开券商APP照着填条件单即可
"""

import os
import datetime
import logging

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl 未安装，请运行: pip install openpyxl")

try:
    import pandas as pd
except ImportError:
    pd = None
    logger.warning("pandas 未安装，请运行: pip install pandas")


def generate_condition_sheet(signals: list, output_dir: str = None,
                             total_capital: float = None) -> str:
    """
    根据今日信号列表生成条件单Excel
    
    参数:
        signals: [(code, signal_dict), ...] 来自 trend_strategy.scan_all_stocks()
        output_dir: 输出目录（默认取config）
        total_capital: 总资金
    
    返回:
        生成的Excel文件路径
    """
    if not HAS_OPENPYXL or pd is None:
        logger.error("缺少 openpyxl 或 pandas，无法生成Excel")
        return ""

    if output_dir is None:
        output_dir = config.OUTPUT_DIR
    if total_capital is None:
        total_capital = config.TOTAL_CAPITAL

    os.makedirs(output_dir, exist_ok=True)
    today = datetime.date.today().strftime("%Y-%m-%d")
    filename = os.path.join(output_dir, f"条件单_{today}.xlsx")

    # 构建数据
    rows = []
    for code, sig in signals:
        stock_info = config.get_stock_info(code)
        name = stock_info.get("名称", code)
        sector = stock_info.get("赛道", "")
        stock_type = stock_info.get("类型", "龙头")

        row = {
            "股票代码": code,
            "股票名称": name,
            "赛道": sector,
            "类型": stock_type,
            "信号类型": _signal_type_cn(sig),
            "买入价": sig.get("buy_price", ""),
            "买入股数": "",
            "初始止损价": sig.get("stop_loss_initial", ""),
            "移动止损价": sig.get("stop_loss_current", ""),
            "第一档止盈": "",
            "第二档止盈": "",
            "信号说明": sig.get("signal_reason", ""),
        }

        # 如果有买入信号，计算仓位
        if sig.get("buy_signal") and sig.get("buy_price"):
            from strategy.position import calc_first_batch
            buy_p = sig["buy_price"]
            stop_p = sig.get("stop_loss_initial", buy_p * 0.9)
            batch = calc_first_batch(buy_p, stop_p, stock_type, total_capital)
            if batch["pass_risk"]:
                row["买入股数"] = batch["shares"]
                # 阶梯止盈价
                row["第一档止盈"] = round(buy_p * (1 + config.LADDER_SELL_LEVELS[0][0]), 2)
                if len(config.LADDER_SELL_LEVELS) > 1:
                    row["第二档止盈"] = round(buy_p * (1 + config.LADDER_SELL_LEVELS[1][0]), 2)
            else:
                row["信号说明"] += f" [风控未过: {batch['risk_msg']}]"

        # 如果有卖出信号，标注卖出价
        if sig.get("sell_signal") and sig.get("sell_price"):
            row["买入价"] = ""
            row["买入股数"] = ""

        rows.append(row)

    # 生成Excel
    df = pd.DataFrame(rows)
    _write_excel(df, filename, today)
    logger.info(f"条件单已生成: {filename}")
    return filename


def _signal_type_cn(sig: dict) -> str:
    """信号类型中文"""
    if sig.get("sell_signal"):
        return "[卖出]"
    if sig.get("buy_signal"):
        return "[买入]"
    if sig.get("add_position"):
        return "[加仓]"
    return "[观望]"


def _write_excel(df: pd.DataFrame, filename: str, today: str):
    """写入带格式的Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"条件单{today}"

    # 标题行
    title_font = Font(name="Microsoft YaHei", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    ws.cell(row=1, column=1, value=f"每日条件单 - {today}").font = title_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # 副标题
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df.columns))
    ws.cell(row=2, column=1,
            value=f"总资金: {config.TOTAL_CAPITAL:,.0f}元 | "
                  f"股票池: {len(config.STOCK_POOL)}只 | "
                  f"生成时间: {datetime.datetime.now().strftime('%H:%M')}").font = Font(size=10)
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 写入表头
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 写入数据
    buy_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    sell_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for row_idx, row_data in df.iterrows():
        excel_row = row_idx + 5
        is_buy = "[买入]" in str(row_data.get("信号类型", ""))
        is_sell = "[卖出]" in str(row_data.get("信号类型", ""))

        for col_idx, col_name in enumerate(df.columns, 1):
            value = row_data[col_name]
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Microsoft YaHei", size=9)

            # 条件格式：买入行绿底，卖出行红底
            if is_buy:
                cell.fill = buy_fill
            elif is_sell:
                cell.fill = sell_fill

            # 数字格式
            if col_name in ("买入价", "初始止损价", "移动止损价", "第一档止盈", "第二档止盈"):
                if isinstance(value, (int, float)) and value > 0:
                    cell.number_format = "#,##0.00"
            elif col_name == "买入股数":
                if isinstance(value, (int, float)) and value > 0:
                    cell.number_format = "#,##0"

    # 列宽
    col_widths = {
        "股票代码": 12, "股票名称": 10, "赛道": 10, "类型": 8,
        "信号类型": 10, "买入价": 10, "买入股数": 10,
        "初始止损价": 12, "移动止损价": 12, "第一档止盈": 12,
        "第二档止盈": 12, "信号说明": 35
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_widths.get(col_name, 12)

    # 底部说明
    note_row = len(df) + 6
    ws.cell(row=note_row, column=1,
            value="说明: 绿色=买入信号, 橙色=卖出信号 | 止损均为收盘价触发 | "
                  "第一档止盈卖出1/3, 第二档止盈再卖1/3").font = Font(size=9, italic=True)

    wb.save(filename)


def generate_simple_report(signals: list) -> str:
    """
    生成纯文本格式的每日信号报告（不依赖openpyxl）
    """
    today = datetime.date.today().strftime("%Y-%m-%d")
    lines = [
        "=" * 60,
        f"  每日交易信号报告 - {today}",
        f"  总资金: {config.TOTAL_CAPITAL:,.0f}元 | 股票池: {len(config.STOCK_POOL)}只",
        "=" * 60,
        ""
    ]

    buy_signals = [(c, s) for c, s in signals if s.get("buy_signal")]
    sell_signals = [(c, s) for c, s in signals if s.get("sell_signal")]
    add_signals = [(c, s) for c, s in signals if s.get("add_position")]
    watch_signals = [(c, s) for c, s in signals if not any([s.get("buy_signal"), s.get("sell_signal"), s.get("add_position")])]

    if sell_signals:
        lines.append(">>> 卖出信号（优先处理）:")
        for code, sig in sell_signals:
            name = config.get_stock_name(code)
            lines.append(f"  {code} {name}: 卖出价 {sig.get('sell_price', '-')} | {sig.get('signal_reason', '')}")
        lines.append("")

    if buy_signals:
        lines.append(">>> 买入信号:")
        for code, sig in buy_signals:
            name = config.get_stock_name(code)
            lines.append(f"  {code} {name}: 买入价 {sig.get('buy_price', '-')} | "
                        f"止损 {sig.get('stop_loss_initial', '-')} | {sig.get('signal_reason', '')}")
        lines.append("")

    if add_signals:
        lines.append(">>> 加仓信号:")
        for code, sig in add_signals:
            name = config.get_stock_name(code)
            lines.append(f"  {code} {name}: {sig.get('signal_reason', '')}")
        lines.append("")

    if watch_signals:
        lines.append(">>> 观望（无信号）:")
        for code, sig in watch_signals:
            name = config.get_stock_name(code)
            lines.append(f"  {code} {name}: {sig.get('signal_reason', '')}")
        lines.append("")

    lines.append("=" * 60)
    return "\n".join(lines)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    print("=" * 50)
    print("  条件单生成模块 - 测试")
    print("=" * 50)

    # 模拟信号
    mock_signals = [
        ("002049", {
            "date": "2025-07-18", "buy_signal": True, "sell_signal": False,
            "buy_price": 200.0, "sell_price": None,
            "stop_loss_initial": 180.0, "stop_loss_current": 180.0,
            "add_position": False, "signal_reason": "缩量回踩20日线: 量比=0.65, MA20=198.5"
        }),
        ("300502", {
            "date": "2025-07-18", "buy_signal": False, "sell_signal": True,
            "buy_price": None, "sell_price": 85.5,
            "stop_loss_initial": None, "stop_loss_current": None,
            "add_position": False, "signal_reason": "回落止盈: 最高92->收盘85.5, 回落7%"
        }),
        ("603986", {
            "date": "2025-07-18", "buy_signal": False, "sell_signal": False,
            "buy_price": None, "sell_price": None,
            "stop_loss_initial": 130.0, "stop_loss_current": 145.0,
            "add_position": True, "signal_reason": "可加仓: 浮盈5.2% >= 3%"
        }),
    ]

    # 文本报告
    report = generate_simple_report(mock_signals)
    print(report)

    # Excel（如果安装了openpyxl）
    if HAS_OPENPYXL:
        path = generate_condition_sheet(mock_signals)
        print(f"\nExcel已生成: {path}")
    else:
        print("\n(未安装openpyxl，跳过Excel生成)")

    print("\n[OK] 条件单模块测试完成")
